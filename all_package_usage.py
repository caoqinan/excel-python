# -*- coding:utf-8 -*-
import xlrd
import xlwt
import openpyxl

xls_filename = r'C:\Users\Administrator\Desktop\mygit\excel-python\1.xls'
xlsx_filename = r'C:\Users\Administrator\Desktop\mygit\excel-python\2.xlsx'

#xlrd, read xls and xlsx
wb = xlrd.open_workbook(xls_filename)
wb.sheet_names() #return sheets name list
wb.sheet_by_index(0) #select sheet by index,start from 0
wb.sheet_by_name() #select sheet by sheet name
sh = wb.sheet_by_index(0)#get one sheet
print sh.name #return name
print sh.nrows
print sh.ncols
#get value
row_value = sh.row_values(0) #return 0 row list
col_value = sh.col_values(0) #return 0 col list
row_value[0],type(row_value[0])
sh.cell(0,0).value,type(sh.cell(0,0).value)
sh.cell(0,0),type(sh.cell(0,0))


#xlwt
workbook = xlwt.Workbook(encoding = 'utf-8')
worksheet = workbook.add_sheet('My Worksheet')
worksheet = workbook.get_sheet(Sheet_name)
worksheet.write(0, 0, 2) # 不带样式的写入
worksheet.write(0, 1, 3, style) # 带样式的写入
worksheet.col(0).width = 4333 # 设置单元格宽度
worksheet.write(0, 2, xlwt.Formula('A1*B1')) # 公式
worksheet.write(0, 3, xlwt.Formula('SUM(A1,B1)')) # 公式
worksheet.write(1, 0, xlwt.Formula('HYPERLINK("http://www.baidu.com");"baidu"'))

style = xlwt.XFStyle() # 初始化样式
font = xlwt.Font() # 为样式创建字体
pattern = xlwt.Pattern() # Create the Pattern
alignment = xlwt.Alignment() # Create Alignment
font.name = 'Times New Roman'
font.bold = True # 黑体
font.colour_index = 2 #red:2 black:0
font.underline = True # 下划线
font.italic = True # 斜体字
alignment.horz = xlwt.Alignment.HORZ_CENTER # May be: HORZ_GENERAL, HORZ_LEFT, HORZ_CENTER, HORZ_RIGHT, HORZ_FILLED, HORZ_JUSTIFIED, HORZ_CENTER_ACROSS_SEL, HORZ_DISTRIBUTED
alignment.vert = xlwt.Alignment.VERT_CENTER # May be: VERT_TOP, VERT_CENTER, VERT_BOTTOM, VERT_JUSTIFIED, VERT_DISTRIBUTED
pattern.pattern = xlwt.Pattern.SOLID_PATTERN # May be: NO_PATTERN, SOLID_PATTERN, or 0x00 through 0x12
pattern.pattern_fore_colour = 5 # May be: 8 through 63. 0 = Black, 1 = White, 2 = Red, 3 = Green, 4 = Blue, 5 = Yellow, 6 = Magenta, 7 = Cyan, 16 = Maroon, 17 = Dark Green, 18 = Dark Blue, 19 = Dark Yellow , almost brown), 20 = Dark Magenta, 21 = Teal, 22 = Light Gray, 23 = Dark Gray, the list goes on...
style.pattern = pattern # Add Pattern to Style
style.alignment = alignment # Add Alignment to Style
style.font = font # 设定样式
worksheet.write_merge(0, 2, 0, 4, 'First Merge',style)#合并单元格 start_row:1,end_row:2,start_col:3,end_col:4
workbook.save('formatting.xls') # 保存文件


#xlrd2xlwt   xlsx 不适用
from xlutils.copy import copy
wb = copy(wb) #xlrd_wb   这个模块是用来在xlrd和xlwt之间嫁接桥梁的，可以说是依赖xlrd的，因为必须要初始化xlrd的Book类才能复制。
sh  = wb.get_sheet(u'第三方支付账户表')
sh.write(6, 6, '2')
wb.save(xls_filename)


#win32com  start_row:1,start_col:1
import win32com.client
xlApp = win32com.client.Dispatch('Excel.Application')
xlBook = xlApp.Workbooks.Open(xls_filename)
# xlBook = xlApp.Workbooks.Add()
xSheet = xlBook.Worksheets(u'第三方支付账户表')
row = col = 3
print xSheet.Cells(row, col).Value #取值
xSheet.Cells(row, col).Value  = '2' #赋值
xSheet.Cells(row, col).Font.Size = 15#字体大小
xSheet.Cells(row, col).Font.Bold = True#是否黑体
xSheet.Cells(row, col).Name = "Arial"#字体类型
xSheet.Cells(row, col).Font.ColorIndex = 3 #字体颜色
xSheet.Cells(row, col).Interior.ColorIndex = 2 #表格背景
#sht.Range("A1").Borders.LineStyle = xlDouble
# xSheet.Cells(row, col).BorderAround(1,4)#表格边框
xSheet.Rows(3).RowHeight = 30#行高
xSheet.Cells(row, col).HorizontalAlignment = -4108 #水平居中xlCenter
xSheet.Cells(row, col).VerticalAlignment = -4108 #垂直
xSheet.Range("A1:A3").Merge() #合并单元格
xSheet.Rows(row).Delete()#删除行
xSheet.Columns(row).Delete()#删除列
xlApp.Worksheets.Add().Name = 'test' #新建test工作表
xSheet1 = xlBook.Worksheets('test')
xSheet.Range('A1').Copy(xSheet1.Range('A1')) #区域复制, xSheet 复制到xSheet1
#复制1
shts = xlBook.Worksheets
shts(1).Copy(None,shts(1))
#复制2
xSheet.Copy(xSheet) #区域复制

columns=xSheet.UsedRange.Columns.Count  #获取最大列
rows = xSheet.UsedRange.Rows.Count    #获取最大行

xlBook.SaveAs(newfilename) #另存为
xlBook.Save() #覆盖保存
xlApp.Quit()#关闭

